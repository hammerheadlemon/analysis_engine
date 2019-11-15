'''
This programme calculates individual project cost profiles.

Output document:
1) excel workbook with a cost profile for each project. Three different cost profiles calculated are i) latest,
ii) last, iii) baseline.

See instructions below.

Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in
the import statement.

TODO: code in need of a refactor. currently adapting so all project outputs go in one wb.

'''

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
from analysis.data import financial_analysis_masters_list, q2_1920, baseline_bc, baseline_index, all_financial_lists, \
    capture_ng, capture_cdel, capture_rdel, capture_income
from analysis.engine_functions import filter_project_group


def financial_data(project_list, masters_list, bl_list, cells_to_capture, index):
    '''
    Function that creates a mini dictionary containing financial information. It used the financial_info function to
    help build the dictionary output.

    project_list: list of project names
    master_list: master data for quarter of interest
    bl_list:
    cells_to_capture: financial info key names. see lists below
    index:

    '''

    output = {}
    for project_name in project_list:
        master_data = masters_list[bl_list[project_name][index]]
        get_financial_info = financial_info(project_name, master_data, cells_to_capture)
        output[project_name] = get_financial_info

    return output

def financial_info(project_name, master_data, cells_to_capture):
    '''
    function that creates dictionary containing financial {key : value} information.

    project_name = name of project
    master_data = quarter master data set
    cells_to_capture = lists of keys of interest
    '''

    output = {}

    if project_name in master_data.projects:
        for item in master_data.data[project_name]:
            if item in cells_to_capture:
                if master_data.data[project_name][item] is None:
                    output[item] = 0
                else:
                    value = master_data.data[project_name][item]
                    output[item] = value

    else:
        for item in cells_to_capture:
            output[item] = 0

    return output

def calculate_totals(project_name, financial_data):
    '''
    Function that calculates totals.

    project_name: project name
    financial_data: mini project financial dictionary
    '''

    working_data = financial_data[project_name]
    rdel_list = []
    cdel_list = []
    ng_list = []

    for rdel in capture_rdel:
        try:
            rdel_list.append(working_data[rdel])
        except KeyError:
            rdel_list.append(int(0))
    for cdel in capture_cdel:
        try:
            cdel_list.append(working_data[cdel])
        except KeyError:
            cdel_list.append(int(0))
    for ng in capture_ng:
        try:
            ng_list.append(working_data[ng])
        except KeyError:
            ng_list.append(int(0))

    total_list = []
    for i in range(len(rdel_list)):
        total = rdel_list[i] + cdel_list[i] + ng_list[i]
        total_list.append(total)

    return total_list

def calculate_income_totals(project_name, financial_data):
    '''
    function that calculates income totals.

    project_name: project name
    financial_data: mini project financial dictionary
    '''

    working_data = financial_data[project_name]
    income_list = []

    for income in capture_income:
        try:
            income_list.append(working_data[income])
        except KeyError:
            income_list.append(int(0))

    return income_list

def place_in_excel_individual(project_name, latest_financial_data, last_financial_data, baseline_financial_data, baseline_record):
    '''
    function places all data into excel spreadsheet and creates chart.
    data is placed into sheet in reverse order (see how data_list is ordered) so that most recent
    data is displayed on right hand side of the data table

    project_name: project name
    latest_financial_data: mini financial dictionary current quarter
    last_financial_data: mini financial dictionary last quarter
    baseline_financial_data: mini financial dictionary baseline quarter

    '''

    wb = Workbook()
    ws = wb.active
    data_list = [baseline_financial_data, last_financial_data, latest_financial_data]
    count = 0

    '''places in raw/reported data'''
    for data in data_list:
        for i, key in enumerate(capture_rdel):
            try:
                ws.cell(row=i+3, column=2+count, value=data[project_name][key])
            except KeyError:
                ws.cell(row=i + 3, column=2 + count, value=0)
        for i, key in enumerate(capture_cdel):
            try:
                ws.cell(row=i+3, column=3+count, value=data[project_name][key])
            except KeyError:
                ws.cell(row=i + 3, column=3 + count, value=0)
        for i, key in enumerate(capture_ng):
            try:
                ws.cell(row=i+3, column=4+count, value=data[project_name][key])
            except KeyError:
                ws.cell(row=i + 3, column=4 + count, value=0)
        count += 4

    '''places in totals'''
    baseline_totals = calculate_totals(project_name, baseline_financial_data)
    last_q_totals = calculate_totals(project_name, last_financial_data)
    latest_q_totals = calculate_totals(project_name, latest_financial_data)

    total_list = [baseline_totals, last_q_totals, latest_q_totals]

    c = 0
    for l in total_list:
        for i, total in enumerate(l):
            ws.cell(row=i+3, column=5+c, value=total)
        c += 4

    '''labeling data in table'''

    labeling_list_quarter = ['Baseline', 'Last Quarter', 'Latest quarter']

    ws.cell(row=1, column=2, value=labeling_list_quarter[0])
    ws.cell(row=1, column=6, value=labeling_list_quarter[1])
    ws.cell(row=1, column=10, value=labeling_list_quarter[2])

    labeling_list_type = ['RDEL', 'CDEL', 'Non-Gov', 'Total']
    repeat = 3
    c = 0
    while repeat > 0:
        for i, label in enumerate(labeling_list_type):
            ws.cell(row=2, column=2+i+c, value=label)
        c += 4
        repeat -= 1

    labeling_list_year = ['Spend', '19/20', '20/21', '21/22', '22/23', '23/24', '24/25', '25/26', '26/27', '27/28',
                          '28/29', 'Unprofiled']

    for i, label in enumerate(labeling_list_year):
        ws.cell(row=2+i, column=1, value=label)

    '''process for showing total cost profile. starting with data'''
    row_start = 16
    for x, l in enumerate(total_list):
        for i, total in enumerate(l):
            ws.cell(row=i + row_start, column=x + 2, value=total)

    '''record which baseline is being used'''
    ws.cell(row=1, column=16).value = 'Baseline quarter'
    ws.cell(row=2, column=16).value = baseline_record[project_name][2][0]

    '''data for graph labeling'''

    for i, quarter in enumerate(labeling_list_quarter):
        ws.cell(row=15, column=i + 2, value=quarter)

    for i, label in enumerate(labeling_list_year):
        ws.cell(row=15+i, column=1, value=label)


    chart = LineChart()
    chart.title = str(project_name) + ' Cost Profile'
    chart.style = 4
    chart.x_axis.title = 'Financial Year'
    chart.y_axis.title = 'Cost £m'

    '''styling chart'''
    # axis titles
    font = Font(typeface='Calibri')
    size = 1200  # 12 point size
    cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chart.x_axis.title.tx.rich.p[0].pPr = pp
    chart.y_axis.title.tx.rich.p[0].pPr = pp
    # chart.title.tx.rich.p[0].pPr = pp

    # title
    size_2 = 1400
    cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
    pp_2 = ParagraphProperties(defRPr=cp_2)
    rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
    chart.title.tx.rich.p[0].pPr = pp_2

    '''unprofiled costs not included in the chart'''
    data = Reference(ws, min_col=2, min_row=15, max_col=4, max_row=25)
    chart.add_data(data, titles_from_data=True)
    cats = Reference(ws, min_col=1, min_row=16, max_row=25)
    chart.set_categories(cats)

    s3 = chart.series[0]
    s3.graphicalProperties.line.solidFill = "cfcfea"  # light blue
    s8 = chart.series[1]
    s8.graphicalProperties.line.solidFill = "5097a4"  # medium blue
    s9 = chart.series[2]
    s9.graphicalProperties.line.solidFill = "0e2f44"  # dark blue'''

    ws.add_chart(chart, "H15")

    '''process for creating income chart'''

    baseline_total_income = calculate_income_totals(project_name, baseline_financial_data)
    last_q_total_income = calculate_income_totals(project_name, last_financial_data)
    latest_q_total_income = calculate_income_totals(project_name, latest_financial_data)

    total_income_list = [baseline_total_income, last_q_total_income, latest_q_total_income]

    if sum(latest_q_total_income) is not 0:
        for x, l in enumerate(total_income_list):
            for i, total in enumerate(l):
                ws.cell(row=i + 32, column=x + 2, value=total)

        '''data for graph labeling'''

        for i, quarter in enumerate(labeling_list_quarter):
            ws.cell(row=32, column=i + 2, value=quarter)

        for i, label in enumerate(labeling_list_year):
            ws.cell(row=32 + i, column=1, value=label)


        '''income graph'''

        chart = LineChart()
        chart.title = str(project_name) + ' Income Profile'
        chart.style = 4
        chart.x_axis.title = 'Financial Year'
        chart.y_axis.title = 'Cost £m'

        font = Font(typeface='Calibri')
        size = 1200  # 12 point size
        cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
        pp = ParagraphProperties(defRPr=cp)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        chart.x_axis.title.tx.rich.p[0].pPr = pp
        chart.y_axis.title.tx.rich.p[0].pPr = pp
        # chart.title.tx.rich.p[0].pPr = pp

        # title
        size_2 = 1400
        cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
        pp_2 = ParagraphProperties(defRPr=cp_2)
        rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
        chart.title.tx.rich.p[0].pPr = pp_2

        #unprofiled costs not included in the chart
        data = Reference(ws, min_col=2, min_row=32, max_col=4, max_row=42)
        chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=33, max_row=42)
        chart.set_categories(cats)


        '''
        keeping as colour coding is useful
        s1 = chart.series[0]
        s1.graphicalProperties.line.solidFill = "cfcfea" #light blue
        s2 = chart.series[1]
        s2.graphicalProperties.line.solidFill = "e2f1bb" #light green 
        s3 = chart.series[2]
        s3.graphicalProperties.line.solidFill = "eaba9d" #light red
        s4 = chart.series[3]
        s4.graphicalProperties.line.solidFil = "5097a4" #medium blue
        s5 = chart.series[4]
        s5.graphicalProperties.line.solidFill = "a0db8e" #medium green
        s6 = chart.series[5]
        s6.graphicalProperties.line.solidFill = "b77575" #medium red
        s7 = chart.series[6]
        s7.graphicalProperties.line.solidFil = "0e2f44" #dark blue
        s8 = chart.series[7]
        s8.graphicalProperties.line.solidFill = "29ab87" #dark green
        s9 = chart.series[8]
        s9.graphicalProperties.line.solidFill = "691c1c" #dark red
        '''

        s3 = chart.series[0]
        s3.graphicalProperties.line.solidFill = "e2f1bb"  # light green
        s8 = chart.series[1]
        s8.graphicalProperties.line.solidFill = "a0db8e"  # medium green
        s9 = chart.series[2]
        s9.graphicalProperties.line.solidFill = "29ab87"  # dark green

        ws.add_chart(chart, "H31")

    else:
        pass

    return wb

def place_in_excel_one_wb(project_name_list, latest_financial_data, last_financial_data, baseline_financial_data, \
                          baseline_record):

    wb = Workbook()

    data_list = [baseline_financial_data, last_financial_data, latest_financial_data]

    for i, project_name in enumerate(project_name_list):
        ws = wb.create_sheet(project_name, i)  # creating worksheets
        ws.title = project_name  # title of worksheet

        '''places in raw/reported data'''
        count = 0
        for data in data_list:
            for i, key in enumerate(capture_rdel):
                try:
                    ws.cell(row=i+3, column=2+count, value=data[project_name][key])
                except KeyError:
                    ws.cell(row=i + 3, column=2 + count, value=0)
            for i, key in enumerate(capture_cdel):
                try:
                    ws.cell(row=i+3, column=3+count, value=data[project_name][key])
                except KeyError:
                    ws.cell(row=i + 3, column=3 + count, value=0)
            for i, key in enumerate(capture_ng):
                try:
                    ws.cell(row=i+3, column=4+count, value=data[project_name][key])
                except KeyError:
                    ws.cell(row=i + 3, column=4 + count, value=0)
            count += 4

        '''places in totals'''
        baseline_totals = calculate_totals(project_name, baseline_financial_data)
        last_q_totals = calculate_totals(project_name, last_financial_data)
        latest_q_totals = calculate_totals(project_name, latest_financial_data)

        total_list = [baseline_totals, last_q_totals, latest_q_totals]

        c = 0
        for l in total_list:
            for i, total in enumerate(l):
                ws.cell(row=i+3, column=5+c, value=total)
            c += 4

        '''labeling data in table'''

        labeling_list_quarter = ['Baseline', 'Last Quarter', 'Latest quarter']

        ws.cell(row=1, column=2, value=labeling_list_quarter[0])
        ws.cell(row=1, column=6, value=labeling_list_quarter[1])
        ws.cell(row=1, column=10, value=labeling_list_quarter[2])

        labeling_list_type = ['RDEL', 'CDEL', 'Non-Gov', 'Total']
        repeat = 3
        c = 0
        while repeat > 0:
            for i, label in enumerate(labeling_list_type):
                ws.cell(row=2, column=2+i+c, value=label)
            c += 4
            repeat -= 1

        labeling_list_year = ['Spend', '19/20', '20/21', '21/22', '22/23', '23/24', '24/25', '25/26', '26/27', '27/28',
                              '28/29', 'Unprofiled']

        for i, label in enumerate(labeling_list_year):
            ws.cell(row=2+i, column=1, value=label)

        '''process for showing total cost profile. starting with data'''
        row_start = 16
        for x, l in enumerate(total_list):
            for i, total in enumerate(l):
                ws.cell(row=i + row_start, column=x + 2, value=total)

        '''record which baseline is being used'''
        ws.cell(row=1, column=16).value = 'Baseline quarter'
        ws.cell(row=2, column=16).value = baseline_record[project_name][2][0]

        '''data for graph labeling'''

        for i, quarter in enumerate(labeling_list_quarter):
            ws.cell(row=15, column=i + 2, value=quarter)

        for i, label in enumerate(labeling_list_year):
            ws.cell(row=15+i, column=1, value=label)


        chart = LineChart()
        chart.title = str(project_name) + ' Cost Profile'
        chart.style = 4
        chart.x_axis.title = 'Financial Year'
        chart.y_axis.title = 'Cost £m'

        '''styling chart'''
        # axis titles
        font = Font(typeface='Calibri')
        size = 1200  # 12 point size
        cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
        pp = ParagraphProperties(defRPr=cp)
        rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
        chart.x_axis.title.tx.rich.p[0].pPr = pp
        chart.y_axis.title.tx.rich.p[0].pPr = pp
        # chart.title.tx.rich.p[0].pPr = pp

        # title
        size_2 = 1400
        cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
        pp_2 = ParagraphProperties(defRPr=cp_2)
        rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
        chart.title.tx.rich.p[0].pPr = pp_2

        '''unprofiled costs not included in the chart'''
        data = Reference(ws, min_col=2, min_row=15, max_col=4, max_row=25)
        chart.add_data(data, titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=16, max_row=25)
        chart.set_categories(cats)

        s3 = chart.series[0]
        s3.graphicalProperties.line.solidFill = "cfcfea"  # light blue
        s8 = chart.series[1]
        s8.graphicalProperties.line.solidFill = "5097a4"  # medium blue
        s9 = chart.series[2]
        s9.graphicalProperties.line.solidFill = "0e2f44"  # dark blue'''

        ws.add_chart(chart, "H15")

        '''process for creating income chart'''

        baseline_total_income = calculate_income_totals(project_name, baseline_financial_data)
        last_q_total_income = calculate_income_totals(project_name, last_financial_data)
        latest_q_total_income = calculate_income_totals(project_name, latest_financial_data)

        total_income_list = [baseline_total_income, last_q_total_income, latest_q_total_income]

        if sum(latest_q_total_income) is not 0:
            for x, l in enumerate(total_income_list):
                for i, total in enumerate(l):
                    ws.cell(row=i + 32, column=x + 2, value=total)

            '''data for graph labeling'''

            for i, quarter in enumerate(labeling_list_quarter):
                ws.cell(row=32, column=i + 2, value=quarter)

            for i, label in enumerate(labeling_list_year):
                ws.cell(row=32 + i, column=1, value=label)


            '''income graph'''

            chart = LineChart()
            chart.title = str(project_name) + ' Income Profile'
            chart.style = 4
            chart.x_axis.title = 'Financial Year'
            chart.y_axis.title = 'Cost £m'

            font = Font(typeface='Calibri')
            size = 1200  # 12 point size
            cp = CharacterProperties(latin=font, sz=size, b=True)  # Bold
            pp = ParagraphProperties(defRPr=cp)
            rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
            chart.x_axis.title.tx.rich.p[0].pPr = pp
            chart.y_axis.title.tx.rich.p[0].pPr = pp
            # chart.title.tx.rich.p[0].pPr = pp

            # title
            size_2 = 1400
            cp_2 = CharacterProperties(latin=font, sz=size_2, b=True)
            pp_2 = ParagraphProperties(defRPr=cp_2)
            rtp_2 = RichText(p=[Paragraph(pPr=pp_2, endParaRPr=cp_2)])
            chart.title.tx.rich.p[0].pPr = pp_2

            #unprofiled costs not included in the chart
            data = Reference(ws, min_col=2, min_row=32, max_col=4, max_row=42)
            chart.add_data(data, titles_from_data=True)
            cats = Reference(ws, min_col=1, min_row=33, max_row=42)
            chart.set_categories(cats)


            '''
            keeping as colour coding is useful
            s1 = chart.series[0]
            s1.graphicalProperties.line.solidFill = "cfcfea" #light blue
            s2 = chart.series[1]
            s2.graphicalProperties.line.solidFill = "e2f1bb" #light green 
            s3 = chart.series[2]
            s3.graphicalProperties.line.solidFill = "eaba9d" #light red
            s4 = chart.series[3]
            s4.graphicalProperties.line.solidFil = "5097a4" #medium blue
            s5 = chart.series[4]
            s5.graphicalProperties.line.solidFill = "a0db8e" #medium green
            s6 = chart.series[5]
            s6.graphicalProperties.line.solidFill = "b77575" #medium red
            s7 = chart.series[6]
            s7.graphicalProperties.line.solidFil = "0e2f44" #dark blue
            s8 = chart.series[7]
            s8.graphicalProperties.line.solidFill = "29ab87" #dark green
            s9 = chart.series[8]
            s9.graphicalProperties.line.solidFill = "691c1c" #dark red
            '''

            s3 = chart.series[0]
            s3.graphicalProperties.line.solidFill = "e2f1bb"  # light green
            s8 = chart.series[1]
            s8.graphicalProperties.line.solidFill = "a0db8e"  # medium green
            s9 = chart.series[2]
            s9.graphicalProperties.line.solidFill = "29ab87"  # dark green

            ws.add_chart(chart, "H31")

        else:
            pass

    return wb


def run_financials_individual(project_name, masters_list):
    '''
    Function that runs this programme.

    project_name: project name
    masters_list: list of master dictionaries

    output 'run' which is in fact a excel wb.
    '''

    '''financial data across different quarters is captured'''
    latest_financial_data = financial_data([project_name], masters_list, q_masters_list, all_data_lists, 0)
    last_financial_data = financial_data([project_name], masters_list, q_masters_list, all_data_lists, 1)
    baseline_financial_data = financial_data([project_name], masters_list, q_masters_list, all_data_lists, 2)

    run = place_in_excel_individual(project_name, latest_financial_data, last_financial_data, baseline_financial_data, baseline_bc)

    return run

def run_financials_one_wb(project_name_list, masters_list):
    '''
    function that runs the programme. It places all project outputs on individual tabs into one wb.
    :param project_name_list: list of project names
    :param masters_list: list containing all quarter master dictionaries
    :return:
    '''

    '''financial data across different quarters is captured'''
    latest_financial_data = financial_data(project_name_list, masters_list, baseline_index, all_financial_lists, 0)
    print(latest_financial_data)
    last_financial_data = financial_data(project_name_list, masters_list, baseline_index, all_financial_lists, 1)
    baseline_financial_data = financial_data(project_name_list, masters_list, baseline_index, all_financial_lists, 2)

    run = place_in_excel_one_wb(project_name_list, latest_financial_data, last_financial_data, baseline_financial_data,
                                baseline_bc)

    return run


'''RUNNING PROGRAMME'''

'''Note: all master data is taken from the data file. Make sure this is up to date and that all relevant data is in 
the import statement.'''

'''ONE. project name list options - this is where the group of interest is specified '''

'''option 1 - all '''
latest_quarter_projects = q2_1920.projects

'''option two - group of projects. use filter_project_group function'''
project_group_list = filter_project_group(q2_1920, 'HSMRPG')

'''option three - single project'''
one_project_list = ['Lower Thames Crossing']


'''TWO. the following for statement prompts the programme to run. 
step one - place the list of projects chosen in step two at the end of the for statement. i.e. for project_name in [here] 
step two - provide relevant file path to document output. Changing the quarter stamp info as necessary. Note keep {} in 
file name as this is where the project name is recorded in the file title'''

# for project_name in one_project_list:
# #     print('Doing financial analysis for ' + str(project_name))
# #     wb = run_financials_single(project_name, financial_analysis_masters_list)
# #     wb.save('C:\\Users\\Standalone\\general\\masters folder\\project_financial_profile\\'
# #             'q2_1920_{}_financial_profile.xlsx'.format(project_name))

output = run_financials_one_wb(latest_quarter_projects, financial_analysis_masters_list)
output.save('C:\\Users\\Standalone\\general\\all_project_financials.xlsx')