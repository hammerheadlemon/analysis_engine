'''

This programme to calculate time difference between reported milestones

input documents:
There quarters master information, typically:
1) latest quarter data
2) last quarter data
3) year ago quarter data

output document:
1) excel workbook with all project milestone information

One thing that this programme does not do is check with milestones have been removed.

See instructions on how to operate programme below.

'''

#TODO solve problem re filtering in excle when values have + sign in front of the them

import datetime
from openpyxl import Workbook
from analysis.engine_functions import all_milestone_data_bulk, ap_p_milestone_data_bulk, \
    assurance_milestone_data_bulk, project_time_difference, bc_ref_stages, get_master_baseline_dict
from analysis.data import q1_1920, list_of_masters_all

def put_into_wb_all(project_name_list, t_data, td_data, td_data_two, wb):
    '''

    Function that places all data into excel wb for this programme

    project_name_list: list of project to return data for
    t_data: dictionary containing milestone data for projects ToDO insert structure
    td_data: dictionary containing time_delta milestone data for projects ToDo insert structure
    td_data_two: dictionary containing second time_delta data for projects ToDO insert structure
    wb: blank excel wb

    '''

    ws = wb.active

    row_num = 2
    for project_name in project_name_list:
        for i, milestone in enumerate(td_data[project_name].keys()):
            ws.cell(row=row_num + i, column=1).value = project_name
            ws.cell(row=row_num + i, column=2).value = milestone
            try:
                milestone_date = tuple(t_data[project_name][milestone])[0]
                ws.cell(row=row_num + i, column=3).value = milestone_date
            except KeyError:
                ws.cell(row=row_num + i, column=3).value = 0

            try:
                value = td_data[project_name][milestone]
                # try:
                    # if int(value) > 0:
                    #     ws.cell(row=row_num + i, column=4).value = '+' + str(value) + ' (days)'
                    # elif int(value) < 0:
                    #     ws.cell(row=row_num + i, column=4).value = str(value) + ' (days)'
                    # elif int(value) == 0:
                ws.cell(row=row_num + i, column=4).value = value
                # except ValueError:
                #     ws.cell(row=row_num + i, column=4).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=4).value = 0

            try:
                value = td_data_two[project_name][milestone]
                # try:
                    # if int(value) > 0:
                    #     ws.cell(row=row_num + i, column=5).value = '+' + str(value) + ' (days)'
                    # elif int(value) < 0:
                    #     ws.cell(row=row_num + i, column=5).value = str(value) + ' (days)'
                    # elif int(value) == 0:
                ws.cell(row=row_num + i, column=5).value = value
                # except ValueError:
                #     ws.cell(row=row_num + i, column=5).value = value
            except KeyError:
                ws.cell(row=row_num + i, column=5).value = 0

            try:
                milestone_date = tuple(t_data[project_name][milestone])[0]
                ws.cell(row=row_num + i, column=6).value = t_data[project_name][milestone][milestone_date]  # provides notes
            except IndexError:
                ws.cell(row=row_num + i, column=6).value = 0

        row_num = row_num + len(td_data[project_name])

    ws.cell(row=1, column=1).value = 'Project'
    ws.cell(row=1, column=2).value = 'Milestone'
    ws.cell(row=1, column=3).value = 'Date'
    ws.cell(row=1, column=4).value = '3/m change (days)'
    ws.cell(row=1, column=5).value = 'Baseline change (days)'
    ws.cell(row=1, column=6).value = 'Notes'

    return wb


def run_milestone_comparator(function, project_name_list, masters_list, date_of_interest):

    '''
    Function that runs this programme.

    function: The type of milestone you wish to analysis can be specified through choosing all_milestone_data_bulk,
    ap_p_milestone_data_bulk, or assurance_milestone_data_bulk functions, all available from engine_function import
    statement above.
    project_name_list: list of project to return data for
    masters_list: list of masters containing quarter information
    date_of_interest: the date after which project milestones should be returned.

    '''


    wb = Workbook()

    '''firstly business cases of interest are filtered out by bc_ref_stage function'''
    baseline_bc = bc_ref_stages(project_name_list, masters_list)
    print(baseline_bc)
    q_masters_list = get_master_baseline_dict(project_name_list, masters_list, baseline_bc)
    print(q_masters_list)

    '''gather mini-dictionaries for each quarter'''

    current_milestones_data = {}
    last_milestones_data = {}
    oldest_milestones_data = {}
    for project_name in project_name_list:
        p_current_milestones_data = function([project_name], masters_list[q_masters_list[project_name][0]])
        current_milestones_data.update(p_current_milestones_data)
        p_last_milestones_data = function([project_name], masters_list[q_masters_list[project_name][1]])
        last_milestones_data.update(p_last_milestones_data)
        p_oldest_milestones_data = function([project_name], masters_list[q_masters_list[project_name][2]])
        oldest_milestones_data.update(p_oldest_milestones_data)

    '''calculate time current and last quarter'''
    first_diff_data = project_time_difference(current_milestones_data, last_milestones_data, date_of_interest)
    second_diff_data = project_time_difference(current_milestones_data, oldest_milestones_data, date_of_interest)

    run = put_into_wb_all(project_name_list, current_milestones_data, first_diff_data, second_diff_data, wb)

    return run


''' RUNNING PROGRAMME '''

'''Note that the all master data is taken from the data file. Make sure that this is up to date and that all relevant
  data is being imported'''

''' ONE. Set relevant list of projects. This needs to be done in accordance with the data you are working with via the
 data.py file '''

'''option one - all projects'''
project_q_list = q1_1920.projects

'''option two - group of projects... in development'''
group_projects_list = []

'''option three - single project'''
one_proj_list = ['Thameslink Programme']

'''TWO. Specify date after which project milestones should be returned. NOTE: Python date format is (YYYY,MM,DD)'''
start_date = datetime.date(2019, 6, 1)

'''THREE. choose the type of variables that you would like to place in run_milestone_comparator function. Arguments 
are placed in this order. 

1. function: The type of milestone you wish to analysis can be specified through choosing all_milestone_data_bulk, 
ap_p_milestone_data_bulk, or assurance_milestone_data_bulk functions, all available from engine_function import 
statement above. 
2. project_name_list: list of project to return data for
3. masters_list: list of masters containing quarter information
4. date_of_interest: the date after which project milestones should be returned. 
 
'''
print_miles = \
    run_milestone_comparator(all_milestone_data_bulk, project_q_list, list_of_masters_all, start_date)

'''FOUR. specify file path to output document'''
print_miles.save('C:\\Users\\Standalone\\general\\testing.xlsx')